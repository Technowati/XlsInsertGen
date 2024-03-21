import datetime
import re
import sys
# # import csv
from tabulate import tabulate
import openpyxl

class ColInfo:
    name = ""
    maxlen = 0

def cleanvalue(vin):
    return str(vin).replace("\n", "~").replace("\r", "~")

multiplierforsize = 1.1   
charstoremove = "ï»¿" 

print("Starting")

if len(sys.argv) <= 3:
    print('No parms!')
    exit()
else:
    fileinname = sys.argv[1]
    fileoutname = sys.argv[2]
    tblname =  sys.argv[3]

outfile = open(fileoutname, "w")

wbin = openpyxl.load_workbook(fileinname)
sheet = wbin.active
cell = sheet.cell(row = 1, column = 1)

numcols = sheet.max_column
numrows = sheet.max_row

for colnum in range(1,numcols+1):
    print(sheet.cell(row=1, column=colnum).value)
print ("numcols", numcols,"numrows", numrows)
print(cell.value) 


# -- Get column info
print("Gathering column info")
cols = []
for c in range(1, numcols+1):
    colname = sheet.cell(row=1, column=c).value.strip().replace(" ", "_")
    print("Column", colname)
    colname = re.sub("[" + charstoremove + "]", "", colname, 0, re.IGNORECASE)
    newcol = ColInfo()
    newcol.name = colname
    newcol.sqlname = "[" + newcol.name + "]"
    newcol.maxlen = 0
    cols.append(newcol)

sqlcolnames = ", ".join(col.sqlname for col in cols)
print("sqlcolnames", sqlcolnames)

# -- Go through once to get stats
print("Gathering file stats")

rownun = 2  # -- Skip headings
cntinsert = 0

for rownum in range(2, numrows+1):
    colvaluesquotedlist = []
    i = -1
    for colnum in range(1, numcols + 1):
        cell = sheet.cell(row=rownum, column=colnum)
        cellvalue = cleanvalue(cell.value)
        colinfo = cols[colnum - 1]

        if (len(cellvalue) > colinfo.maxlen):
            colinfo.maxlen = len(cellvalue)

outfile.write("-- == CsvInsertGen.py == -- \n")
outfile.write("-- Created       = {datetime}\n".format(datetime=datetime.datetime.now()))
outfile.write("-- fileinname    = {fileinname}\n".format(fileinname=fileinname))
outfile.write("-- fileoutname   = {fileoutname}\n".format(fileoutname=fileoutname))
outfile.write("-- tblname       = {tblname}\n".format(tblname=tblname))
outfile.write("-- Total rows in = {numrows:,} (+ 1 header)\n".format(numrows=numrows))
outfile.write("\n")

coltab = []
for colinfo in cols:
    coltab.append([colinfo.name, colinfo.maxlen])

# -- Generate the DDL

outfile.write("SET NOCOUNT ON\n")
outfile.write("\n")

outfile.write("IF EXISTS(Select * from sys.objects Where Name = '{tblname}' and Type = 'U')\n".format(tblname=tblname))
outfile.write("BEGIN\n")
outfile.write("    DROP TABLE {tblname}\n".format(tblname=tblname))
outfile.write("END\n")
outfile.write("GO\n")
outfile.write("\n")

outfile.write("CREATE TABLE {tblname}\n".format(tblname=tblname))
outfile.write("(\n")
outfile.write("  _ID INT NOT NULL IDENTITY (1,1)\n")

for colinfo in cols:
    # coltab.append([colinfo.name, colinfo.maxlen])
    collentouse = int(max([50, colinfo.maxlen * multiplierforsize]))

    outfile.write(", {colname} NVARCHAR({collentouse}) NULL".format(colname=colinfo.sqlname, collentouse=collentouse))
    outfile.write("\n".format(colname=colinfo.name))

outfile.write(")\n")
outfile.write("GO\n")
outfile.write("\n")

# -- Now go through to do actual work

print("Processing file")

for rownum in range(2, numrows+1):

    # ### DANGER! Use this code only when selecting/running for individual providers.
    # selectedproviders = [17924, 17926, 19427, 20154, 7565]
    # if (sheet.cell(row=rownum, column=1).value not in selectedproviders):
    #     continue

    colvaluesquotedlist = []
    for colnum in range(1, numcols + 1):
        cell = sheet.cell(row=rownum, column=colnum)
        cellvalue = cleanvalue(cell.value)
        colinfo = cols[colnum - 1]
        colvaluequoted = "Null" if cellvalue == "None" else "\'" + (cellvalue).replace("\'","\'\'") + "\'"
        colvaluesquotedlist.append(colvaluequoted)

    cntinsert = cntinsert + 1
    outfile.write("INSERT {tblname} ({sqlcolnames})\n".format(tblname=tblname, sqlcolnames=sqlcolnames))
    colvaluesquoted = ", ".join(colvaluesquotedlist)
    colvaluesquoted = colvaluesquoted.replace("\u2002", " ")  # Special 
    outfile.write("VALUES ({colvaluesquoted})\n".format(colvaluesquoted=colvaluesquoted))
    outfile.write("\n")    

    if ((rownum % 1000) == 0):
        outfile.write("GO\n")    
        outfile.write("\n")    
        outfile.write("print 'Finished #{rownum}'\n".format(rownum=rownum))    
        outfile.write("\n")    

outfile.write("SELECT Cnt = COUNT(*) FROM {tblname}\n".format(tblname=tblname))
outfile.write("GO\n")    
outfile.write("\n")

outfile.write("-- Total to insert {cntinsert:,}\n".format(cntinsert=cntinsert))

outfile.close()

print()
print("File In:", fileinname)    
print("File Out:", fileoutname)    
print("Rows:", numrows)    
print("Inserts:", cntinsert)    
print()    
print(tabulate(coltab, headers=["name","maxlen"]))
print()

# print ("Exiting, not really doing the rest for now...")
# exit()  # -- Not really doing the rest until modified
